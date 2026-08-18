"""配置编辑器 — Excel → JSON 生成器

把 config/配置编辑器.xlsx 中的配置写回对应的 JSON 配置文件。

当前支持的 sheet：
    - 时间配置：更新 cleaning_config.json 的「时间范围」部分

用法:
    python scripts/config_excel_to_json.py            # 生成并写回
    python scripts/config_excel_to_json.py --init     # 首次创建 Excel 模板（从当前 JSON 导出）
    python scripts/config_excel_to_json.py --dry-run  # 只读不改写（打印将生成的内容）

设计原则：
    - Excel 是「编辑层」，JSON 是「事实源」（系统只读 JSON）
    - 生成器只更新 JSON 中对应小节，不触碰其他配置
    - 校验失败时拒绝写回，避免脏数据进入系统
"""
from __future__ import annotations

import json
import sys
from datetime import datetime
from pathlib import Path

try:
    import openpyxl
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("缺少 openpyxl，请先安装: pip install openpyxl")
    sys.exit(1)

BASE_DIR = Path(__file__).parent.parent
EXCEL_PATH = BASE_DIR / "config" / "配置编辑器.xlsx"
CLEANING_CFG = BASE_DIR / "config" / "清洗配置" / "cleaning_config.json"
DISPLAY_RULES_CFG = BASE_DIR / "config" / "前端渲染" / "展示规则.json"
ATTRIBUTION_CFG = BASE_DIR / "config" / "清洗配置" / "客户销售归属.json"

# 支持的时间配置键 → 说明
TIME_KEYS = ["年度累计", "月度数据", "季度累计筛选", "年基线数据"]

# dynamic 策略白名单（与 engine/core/config.py _STRATEGY_REGISTRY 对应）
DYNAMIC_STRATEGIES = ["last_full_month", "last_full_quarter"]

HEADERS = ["配置名", "模式", "动态策略", "开始日期", "结束日期", "年份", "月份范围", "说明"]

# 展示规则 sheet 列
# 路径 = 从页面下到叶子的点路径，如：
#   销售TopN                    页面下标量
#   客户矩阵.最大行数             区块.标量
#   部门卡.显示                  区块.布尔
#   客户矩阵.优先展示.1           区块.数组.序号（数组项每行带序号）
#   客户矩阵.客户筛选             空数组用「值」留空表示
RULE_HEADERS = ["页面", "路径", "值", "说明"]

# 展示规则：顶层页面键（固定顺序）
RULE_PAGES = ["数据总览", "年度达成", "月度达成", "季度达成", "销售达成", "年度同比"]

# 数组型配置项：路径中以「路径.序号」形式出现 → JSON 数组
ARRAY_KEYS = {"优先展示", "客户筛选"}

# 页面 → 区块 层级定义（用于模板初始化时按序输出）
# 值为区块名列表；空 [] 表示无区块（配置项直接在页面下）
RULE_SECTIONS = {
    "数据总览": [],
    "年度达成": ["部门卡", "客户矩阵"],
    "月度达成": ["部门卡", "客户矩阵"],
    "季度达成": ["部门卡", "客户矩阵"],
    "销售达成": [],
    "年度同比": [],
}

# 布尔型配置项（Excel 值 true/false → JSON bool）
BOOL_KEYS = {"显示"}

# 销售归属 sheet 列
ATT_HEADERS = ["母公司", "子公司", "指标", "部门", "销售", "比例", "说明"]

# 指标取值（收入,回款 表示同时写入两个指标）
METRIC_VALUES = ["收入", "回款", "收入,回款"]

# 部门取值
DEPT_VALUES = ["检测", "信息", "能源", "海外"]


# ──────────────────────────────────────────────────────────────
# Excel 读取 → 时间配置 dict
# ──────────────────────────────────────────────────────────────
def _parse_cell(value) -> str:
    """单元格值 → 干净字符串（Excel 日期/数字转文本）"""
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d")
    return str(value).strip()


def _parse_month_range(raw: str) -> list[int]:
    """解析月份范围列：'1-8' / '1,8' → [1, 8]"""
    raw = raw.strip().strip("[]").replace("，", ",")
    if not raw:
        return []
    parts = [p.strip() for p in raw.replace("-", ",").split(",") if p.strip()]
    nums = []
    for p in parts:
        try:
            nums.append(int(p))
        except ValueError:
            raise ValueError(f"月份范围格式错误: {raw!r}（应为 '1-8' 或 '1,8'）")
    return nums


def excel_to_time_config(sheet) -> dict:
    """从 Excel「时间配置」sheet 构建 {键: 配置dict}"""
    result: dict = {}
    seen: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name = _parse_cell(row[0])
        if not name:
            continue  # 空行跳过
        if name in seen:
            raise ValueError(f"时间配置存在重复键: {name}")
        seen.add(name)

        mode = _parse_cell(row[1]) or "static"
        strategy = _parse_cell(row[2])
        start = _parse_cell(row[3])
        end = _parse_cell(row[4])
        year = _parse_cell(row[5])
        month_range = _parse_cell(row[6])
        note = _parse_cell(row[7])

        if mode == "dynamic":
            if strategy not in DYNAMIC_STRATEGIES:
                raise ValueError(
                    f"{name}: 未知动态策略 {strategy!r}，可选 {DYNAMIC_STRATEGIES}"
                )
            cfg: dict = {"_mode": "dynamic", "_strategy": strategy}
        elif mode == "static":
            if name == "年基线数据":
                # 年基线：年份 + 月份范围
                if not year:
                    raise ValueError(f"{name}: static 模式需填「年份」")
                months = _parse_month_range(month_range)
                if len(months) != 2 or months[0] > months[1]:
                    raise ValueError(f"{name}: 月份范围应为 [起始月, 结束月]，如 1-8")
                cfg = {"年份": int(year), "月份范围": months}
            else:
                # 常规时间范围：开始/结束日期
                if not start or not end:
                    raise ValueError(f"{name}: static 模式需填「开始日期」和「结束日期」")
                for v in (start, end):
                    try:
                        datetime.strptime(v, "%Y-%m-%d %H:%M:%S")
                    except ValueError:
                        try:
                            datetime.strptime(v, "%Y-%m-%d")
                        except ValueError:
                            raise ValueError(f"{name}: 日期格式错误 {v!r}（应为 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS）")
                cfg = {"start_date": start, "end_date": end}
        else:
            raise ValueError(f"{name}: 未知模式 {mode!r}（应为 static 或 dynamic）")

        if note:
            cfg["_使用方"] = note
        result[name] = cfg

    # 校验必备键齐全
    for k in TIME_KEYS:
        if k not in result:
            raise ValueError(f"Excel 时间配置缺少必备项: {k}")

    return result


# ──────────────────────────────────────────────────────────────
# Excel 读取 → 展示规则 dict
# ──────────────────────────────────────────────────────────────
def _type_cast(value: str):
    """按字符串内容推断值类型：bool / int / float / str"""
    v = value.strip()
    low = v.lower()
    if low in ("true", "false"):
        return low == "true"
    try:
        return int(v)
    except ValueError:
        pass
    try:
        return float(v)
    except ValueError:
        pass
    return v


def _get_path_node(root: dict, parent_parts: list[str]) -> dict:
    """沿路径深入 root，返回 parent_parts 处的容器 dict。

    Args:
        root: 页面 dict
        parent_parts: 键之前的完整父路径（如 ["部门卡"] 表示在 部门卡 下写键）
    """
    node = root
    for p in parent_parts:
        if p not in node or not isinstance(node[p], dict):
            node[p] = {}
        node = node[p]
    return node


def excel_to_display_rules(sheet) -> dict:
    """从 Excel「展示规则」sheet 构建展示规则 dict（路径列结构）

    路径规则:
        - "销售TopN"                  → 页面下 标量
        - "客户矩阵.最大行数"           → 区块.标量
        - "部门卡.显示"                → 区块.布尔
        - "客户矩阵.优先展示"           → 区块.数组（空数组：值留空，无序号行）
        - "客户矩阵.优先展示.1"         → 区块.数组.序号
    """
    result: dict = {}
    # 暂存数组项: {page: {section: {arr_key: [(seq, value)]}}}
    arrays: dict[str, dict[str, dict]] = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        page = _parse_cell(row[0])
        path = _parse_cell(row[1])
        value = _parse_cell(row[2])
        note = _parse_cell(row[3])
        if not page or not path:
            continue

        parts = path.split(".")
        # 数组项：路径最后一段是数字 且 倒数第二段是 ARRAY_KEYS
        if len(parts) >= 2 and parts[-1].isdigit() and parts[-2] in ARRAY_KEYS:
            section = ".".join(parts[:-2])
            arr_key = parts[-2]
            seq = int(parts[-1])
            arrays.setdefault(page, {}).setdefault(section, {}).setdefault(arr_key, []).append((seq, value))
            continue

        # 判断是否数组键（无序号行 = 空数组占位 或 数组无项）
        if parts[-1] in ARRAY_KEYS:
            section = ".".join(parts[:-1])
            # 值留空 → 空数组；否则后续有序号行
            arrays.setdefault(page, {}).setdefault(section, {})[parts[-1]] = []
            continue

        # 普通标量/布尔：写入嵌套
        section = ".".join(parts[:-1])
        key = parts[-1]
        node = _get_path_node(result.setdefault(page, {}), parts[:-1])
        node[key] = _type_cast(value)
        if note:
            node[f"_{key}说明"] = note

    # 合并数组项
    for page, sections in arrays.items():
        page_obj = result.setdefault(page, {})
        for section, arr_map in sections.items():
            node = _get_path_node(page_obj, section.split(".")) if section else page_obj
            for arr_key, items in arr_map.items():
                items_sorted = sorted(items, key=lambda x: x[0])
                node[arr_key] = [v for _, v in items_sorted]

    # 校验必备页面
    for p in RULE_PAGES:
        if p not in result:
            raise ValueError(f"展示规则缺少页面: {p}")

    return result


# ──────────────────────────────────────────────────────────────
# 生成器主流程
# ──────────────────────────────────────────────────────────────
def _render_value(v, depth: int) -> str:
    """递归渲染 JSON 值，复刻原文件风格：
    - dict → 多行 + 2空格缩进
    - list（如月份范围）→ 紧凑单行
    - depth: 值内容所用的缩进层级（depth=2 → 内容缩进4空格）
    """
    if isinstance(v, dict):
        if not v:
            return "{}"
        lines = ["{"]
        items = list(v.items())
        for i, (k, val) in enumerate(items):
            sep = "," if i < len(items) - 1 else ""
            lines.append(f'{"  " * depth}"{k}": {_render_value(val, depth + 1)}{sep}')
        lines.append(f'{"  " * (depth - 1)}}}')
        return "\n".join(lines)
    if isinstance(v, list):
        return "[" + ", ".join(json.dumps(x, ensure_ascii=False) for x in v) + "]"
    return json.dumps(v, ensure_ascii=False)


def _render_time_range_block(time_config: dict, start_indent: int = 2) -> str:
    """渲染「时间范围」块文本，复刻原文件风格（键间空行 + 正确缩进）。

    Args:
        time_config: 时间范围 dict
        start_indent: "时间范围" 键的缩进空格数（顶层键 = 2）
    """
    base = " " * start_indent
    out: list[str] = [f'{base}"时间范围": {{']
    keys = list(time_config.items())
    for i, (k, v) in enumerate(keys):
        rendered = _render_value(v, (start_indent // 2) + 2)
        lines = rendered.splitlines()
        sep = "," if i < len(keys) - 1 else ""
        # 顶层键（如 "年度累计"）缩进 base + 2 空格
        out.append(f'{base}  "{k}": {lines[0]}')
        if len(lines) > 1:
            out.extend(lines[1:])
        # 逗号加在值结束之后（多行 dict → 最后一行；单行 → 键行）
        out[-1] += sep
        if i < len(keys) - 1:
            out.append("")  # 键间空行
    out.append(f"{base}}}")
    return "\n".join(out)


def _replace_time_range_block(text: str, time_config: dict) -> str:
    """在 JSON 文本中精确替换「时间范围」块（从键所在行行首 到 匹配闭合的 }），其余字节不动"""
    start = text.find('"时间范围"')
    if start == -1:
        raise ValueError("cleaning_config.json 中找不到「时间范围」键")
    # 定位到该键所在行行首（保留前导缩进与换行）
    line_start = text.rfind("\n", 0, start) + 1
    # 找到冒号后的第一个 {
    brace = text.find("{", start)
    # 括号深度匹配到闭合 }
    depth = 0
    in_str = False
    escape = False
    end = None
    for i in range(brace, len(text)):
        ch = text[i]
        if in_str:
            if escape:
                escape = False
            elif ch == "\\":
                escape = True
            elif ch == '"':
                in_str = False
        else:
            if ch == '"':
                in_str = True
            elif ch == "{":
                depth += 1
            elif ch == "}":
                depth -= 1
                if depth == 0:
                    end = i
                    break
    if end is None:
        raise ValueError("cleaning_config.json 中「时间范围」块未正确闭合")

    new_block = _render_time_range_block(time_config, start_indent=2)
    # 消费原块后的逗号（若存在）
    tail = text[end + 1:]
    if tail.lstrip().startswith(","):
        consumed = tail.lstrip()[1:]
        new_block += ","
    else:
        consumed = tail
    return text[:line_start] + new_block + consumed


def update_cleaning_config(time_config: dict, dry_run: bool = False) -> dict:
    """把时间配置写回 cleaning_config.json 的「时间范围」部分（文本级替换，最小 diff）"""
    if not CLEANING_CFG.exists():
        raise FileNotFoundError(f"cleaning_config.json 不存在: {CLEANING_CFG}")

    text = CLEANING_CFG.read_text(encoding="utf-8")
    cfg = json.loads(text)
    old_tr = cfg.get("时间范围", {})
    # 保留 _说明（顶层 + 年基线数据）
    if "_说明" in old_tr:
        time_config = {"_说明": old_tr["_说明"], **time_config}
    if old_tr.get("年基线数据", {}).get("_说明"):
        if "年基线数据" in time_config:
            time_config["年基线数据"] = {
                "_说明": old_tr["年基线数据"]["_说明"],
                **time_config["年基线数据"],
            }

    if dry_run:
        return {**cfg, "时间范围": time_config}

    new_text = _replace_time_range_block(text, time_config)
    CLEANING_CFG.write_text(new_text, encoding="utf-8")
    return json.loads(new_text)


# ──────────────────────────────────────────────────────────────
# 展示规则写回（保留头部说明字段，页面数据用标准格式重建）
# ──────────────────────────────────────────────────────────────
def _merge_notes(merged: dict, old: dict):
    """递归合并 _ 前缀说明字段：merged 为 Excel 结果，old 为原文件页面

    作用：Excel 不维护 _说明 等注释字段，写回时把原文件的说明补回，避免信息丢失。
    """
    for k, v in old.items():
        if isinstance(k, str) and k.startswith("_"):
            merged.setdefault(k, v)
        elif isinstance(v, dict) and isinstance(merged.get(k), dict):
            _merge_notes(merged[k], v)


def update_display_rules(rules: dict, dry_run: bool = False) -> dict:
    """把展示规则写回展示规则.json（保留 _说明/_关键约定，重建 6 个页面部分）

    设计：不追求字节级复刻原文件排版，而是保证：
        1. JSON 永远有效（用 json.dumps 标准输出）
        2. _ 前缀说明字段（_说明/_关键约定/区块 _说明 等）从原文件保留
        3. 页面数据 100% 来自 Excel
    """
    if not DISPLAY_RULES_CFG.exists():
        raise FileNotFoundError(f"展示规则.json 不存在: {DISPLAY_RULES_CFG}")

    text = DISPLAY_RULES_CFG.read_text(encoding="utf-8")
    cfg = json.loads(text)
    new_cfg = dict(cfg)  # 保留 _说明/_关键约定 及一切现有字段
    # 页面数据替换为 Excel 值；同时保留原页面中 _ 前缀说明字段（_客户筛选说明/_说明 等）
    for page in RULE_PAGES:
        excel_page = rules[page]
        old_page = cfg.get(page, {})
        merged = dict(excel_page)
        _merge_notes(merged, old_page)
        new_cfg[page] = merged

    if dry_run:
        return new_cfg

    # 标准格式输出（ensure_ascii=False 保留中文；indent=2；末尾换行）
    DISPLAY_RULES_CFG.write_text(
        json.dumps(new_cfg, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return new_cfg


# ──────────────────────────────────────────────────────────────
# Excel 读取 → 销售归属 dict
# ──────────────────────────────────────────────────────────────
def excel_to_attribution(sheet) -> dict:
    """从 Excel「销售归属」sheet 构建 {母公司: {子公司: {指标: {部门: {销售: 比例}}}}}

    指标列取值：收入 / 回款 / 收入,回款（同时写入两个指标）
    """
    # 读取行
    rows: list[dict] = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        parent = _parse_cell(row[0])
        sub = _parse_cell(row[1])
        metric = _parse_cell(row[2])
        dept = _parse_cell(row[3])
        sales = _parse_cell(row[4])
        ratio = _parse_cell(row[5])
        note = _parse_cell(row[6])
        if not parent and not sub:
            continue  # 空行
        if not parent or not sub or not metric or not dept or not sales:
            raise ValueError(
                f"销售归属存在不完整行（母公司/子公司/指标/部门/销售 必填）: {parent}/{sub}/{metric}/{dept}/{sales}"
            )
        rows.append({
            "parent": parent, "sub": sub, "metric": metric,
            "dept": dept, "sales": sales, "ratio": ratio, "note": note,
        })

    if not rows:
        raise ValueError("销售归属 sheet 没有数据行")

    # 构建嵌套结构
    result: dict = {}
    for r in rows:
        metrics = [m.strip() for m in r["metric"].split(",") if m.strip()]
        for m in metrics:
            if m not in ("收入", "回款"):
                raise ValueError(f"非法指标值: {m!r}（可选 收入/回款/收入,回款）")
        # 比例解析（支持整数/小数/百分比/空=1.0）
        ratio = r["ratio"]
        if ratio == "":
            ratio_v = 1.0
        elif ratio.endswith("%"):
            ratio_v = float(ratio[:-1]) / 100.0
        else:
            ratio_v = float(ratio)
        if not (0 < ratio_v <= 1.0):
            raise ValueError(f"比例非法（应为 0~1）: {r['parent']}/{r['sub']}/{r['sales']} = {ratio_v}")

        parent_obj = result.setdefault(r["parent"], {"子公司": {}})
        subs = parent_obj.setdefault("子公司", {})
        sub_obj = subs.setdefault(r["sub"], {})
        for m in metrics:
            dept_map = sub_obj.setdefault(m, {}).setdefault(r["dept"], {})
            dept_map[r["sales"]] = ratio_v
            if r["note"]:
                sub_obj.setdefault("_说明", r["note"])

    return result


def update_attribution(attribution: dict, dry_run: bool = False) -> dict:
    """把销售归属写回 客户销售归属.json（保留 _说明，客户归属数据来自 Excel）"""
    if not ATTRIBUTION_CFG.exists():
        raise FileNotFoundError(f"客户销售归属.json 不存在: {ATTRIBUTION_CFG}")

    text = ATTRIBUTION_CFG.read_text(encoding="utf-8")
    cfg = json.loads(text)
    new_cfg = {"_说明": cfg.get("_说明", "客户统一归属 — 母公司→子公司→销售分配"),
               "客户归属": attribution}

    if dry_run:
        return new_cfg

    ATTRIBUTION_CFG.write_text(
        json.dumps(new_cfg, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    return new_cfg


# ──────────────────────────────────────────────────────────────
# Excel 模板初始化（从当前 JSON 导出）
# ──────────────────────────────────────────────────────────────
def _cell_fill(name: str) -> str:
    fills = {
        "年度累计": "DDEBF7",
        "月度数据": "E2EFDA",
        "季度累计筛选": "FFF2CC",
        "年基线数据": "FCE4D6",
    }
    return fills.get(name, "FFFFFF")


def _append_rule_row(ws, row_idx: int, page: str, path: str, value, note: str, is_bool: bool = False):
    """写一行展示规则（路径列结构）

    Args:
        page: 页面名
        path: 路径（如 "客户矩阵.最大行数" / "客户矩阵.优先展示.1"）
        value: 值（数组项为 str；空数组传 None；bool 传 bool）
        note: 说明
        is_bool: 值是否为布尔
    """
    thin = Side(style="thin", color="B0B7C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    display = str(value) if value is not None else ""
    if is_bool:
        display = "true" if value else "false"
    vals = [page, path, display, note]
    for c, v in enumerate(vals, 1):
        cell = ws.cell(row=row_idx, column=c, value=v)
        cell.border = border
        cell.alignment = Alignment(vertical="center", wrap_text=(c == 4))


def init_excel_template():
    """从 cleaning_config.json 当前时间范围导出 Excel 模板"""
    cfg = json.loads(CLEANING_CFG.read_text(encoding="utf-8"))
    tr = cfg.get("时间范围", {})

    wb = openpyxl.Workbook()

    # ── Sheet 1: 时间配置 ──
    ws = wb.active
    ws.title = "时间配置"
    ws.append(HEADERS)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    thin = Side(style="thin", color="B0B7C3")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for col, h in enumerate(HEADERS, 1):
        c = ws.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    for key in TIME_KEYS:
        spec = tr.get(key, {})
        fill_color = _cell_fill(key)
        if spec.get("_mode") == "dynamic":
            values = [key, "dynamic", spec.get("_strategy", ""), "", "", "", "", spec.get("_使用方", "")]
        elif key == "年基线数据":
            months = spec.get("月份范围", [])
            values = [
                key, "static", "",
                "", "", spec.get("年份", ""),
                f"{months[0]}-{months[1]}" if len(months) == 2 else "",
                spec.get("_使用方", ""),
            ]
        else:
            values = [
                key, "static", "",
                spec.get("start_date", ""), spec.get("end_date", ""),
                "", "", spec.get("_使用方", ""),
            ]
        ws.append(values)
        r = ws.max_row
        for col in range(1, len(HEADERS) + 1):
            c = ws.cell(row=r, column=col)
            c.border = border
            c.fill = PatternFill("solid", fgColor=fill_color)
            c.alignment = Alignment(vertical="center", wrap_text=(col == 8))

    # 列宽
    widths = [16, 12, 20, 18, 26, 8, 12, 40]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # 数据验证：模式列下拉
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_mode = DataValidation(type="list", formula1='"static,dynamic"', allow_blank=True)
    dv_mode.add(f"B2:B{ws.max_row + 50}")
    ws.add_data_validation(dv_mode)
    dv_strat = DataValidation(type="list", formula1='"last_full_month,last_full_quarter"', allow_blank=True)
    dv_strat.add(f"C2:C{ws.max_row + 50}")
    ws.add_data_validation(dv_strat)

    ws.freeze_panes = "A2"

    # ── Sheet 2: 展示规则 ──
    rules = json.loads(DISPLAY_RULES_CFG.read_text(encoding="utf-8"))
    ws_r = wb.create_sheet("展示规则")
    ws_r.append(RULE_HEADERS)
    for col, h in enumerate(RULE_HEADERS, 1):
        c = ws_r.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    row_idx = 2
    for page in RULE_PAGES:
        page_obj = rules.get(page, {})
        sections = RULE_SECTIONS.get(page, [])
        if sections:
            # 有区块的页面：先顶层标量键，再各区块
            top_keys = [k for k in page_obj.keys()
                        if not isinstance(page_obj[k], dict) and not k.startswith("_")]
            for key in top_keys:
                _append_rule_row(ws_r, row_idx, page, key, page_obj[key],
                                 page_obj.get(f"_{key}说明", ""), is_bool=key in BOOL_KEYS)
                row_idx += 1
            for sec in sections:
                sec_obj = page_obj.get(sec, {})
                if not isinstance(sec_obj, dict):
                    continue
                for key, val in sec_obj.items():
                    if key.startswith("_"):
                        continue
                    path = f"{sec}.{key}"
                    note = sec_obj.get(f"_{key}说明", "")
                    if key in ARRAY_KEYS and isinstance(val, list):
                        if not val:
                            # 空数组：占位一行，值留空
                            _append_rule_row(ws_r, row_idx, page, path, None, note)
                            row_idx += 1
                        else:
                            for i, v in enumerate(val, 1):
                                _append_rule_row(ws_r, row_idx, page, f"{path}.{i}", v, note)
                                row_idx += 1
                    else:
                        _append_rule_row(ws_r, row_idx, page, path, val, note,
                                         is_bool=key in BOOL_KEYS)
                        row_idx += 1
        else:
            # 无区块页面：全部键在页面下（含顶层数组/布尔/内嵌 dict）
            for key, val in page_obj.items():
                if key.startswith("_"):
                    continue
                note = page_obj.get(f"_{key}说明", "")
                if isinstance(val, dict) and val:
                    # 内嵌 dict：子路径展开（如 卡片1_销售达成.显示）
                    for sub_key, sub_val in val.items():
                        _append_rule_row(ws_r, row_idx, page, f"{key}.{sub_key}",
                                         sub_val, note, is_bool=sub_key in BOOL_KEYS)
                        row_idx += 1
                elif key in ARRAY_KEYS and isinstance(val, list):
                    if not val:
                        _append_rule_row(ws_r, row_idx, page, key, None, note)
                        row_idx += 1
                    else:
                        for i, v in enumerate(val, 1):
                            _append_rule_row(ws_r, row_idx, page, f"{key}.{i}", v, note)
                            row_idx += 1
                else:
                    _append_rule_row(ws_r, row_idx, page, key, val,
                                     note, is_bool=key in BOOL_KEYS)
                    row_idx += 1
        if row_idx > 2:
            row_idx += 1  # 页面间空行

    # 列宽
    r_widths = [16, 40, 40, 44]
    for i, w in enumerate(r_widths, 1):
        ws_r.column_dimensions[get_column_letter(i)].width = w

    ws_r.freeze_panes = "A2"

    # ── Sheet 3: 销售归属 ──
    att = json.loads(ATTRIBUTION_CFG.read_text(encoding="utf-8"))["客户归属"]
    ws_a = wb.create_sheet("销售归属")
    ws_a.append(ATT_HEADERS)
    for col, h in enumerate(ATT_HEADERS, 1):
        c = ws_a.cell(row=1, column=col)
        c.fill = header_fill
        c.font = header_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border

    att_row = 2
    att_fill = PatternFill("solid", fgColor="FFFFFF")
    for parent, group in att.items():
        for sub, sub_data in group.get("子公司", {}).items():
            # 判断收入/回款是否同构：同构 → 指标=收入,回款 一行
            inc = sub_data.get("收入", {})
            pay = sub_data.get("回款", {})
            if inc == pay and inc:
                note = sub_data.get("_说明", "")
                for dept, sales_map in inc.items():
                    for sales, ratio in sales_map.items():
                        vals = [parent, sub, "收入,回款", dept, sales,
                                f"{ratio:.15g}", note]
                        for c, v in enumerate(vals, 1):
                            cell = ws_a.cell(row=att_row, column=c, value=v)
                            cell.border = border
                            cell.fill = att_fill
                            cell.alignment = Alignment(vertical="center")
                        att_row += 1
            else:
                # 结构不同：收入/回款分开行
                for metric, m in (("收入", inc), ("回款", pay)):
                    if not m:
                        continue
                    note = sub_data.get("_说明", "")
                    for dept, sales_map in m.items():
                        for sales, ratio in sales_map.items():
                            vals = [parent, sub, metric, dept, sales,
                                    f"{ratio:.15g}", note]
                            for c, v in enumerate(vals, 1):
                                cell = ws_a.cell(row=att_row, column=c, value=v)
                                cell.border = border
                                cell.fill = att_fill
                                cell.alignment = Alignment(vertical="center")
                            att_row += 1
        if att_row > 2:
            att_row += 1  # 母公司间空行

    # 列宽
    a_widths = [30, 40, 14, 10, 12, 10, 40]
    for i, w in enumerate(a_widths, 1):
        ws_a.column_dimensions[get_column_letter(i)].width = w

    # 数据验证：指标 / 部门 / 比例
    from openpyxl.worksheet.datavalidation import DataValidation
    dv_metric = DataValidation(type="list", formula1='"收入,回款,收入,回款"', allow_blank=True)
    dv_metric.add(f"C2:C{att_row + 50}")
    ws_a.add_data_validation(dv_metric)
    dv_dept = DataValidation(type="list", formula1='"检测,信息,能源,海外"', allow_blank=True)
    dv_dept.add(f"D2:D{att_row + 50}")
    ws_a.add_data_validation(dv_dept)

    ws_a.freeze_panes = "A2"

    # ── Sheet 4: 说明 ──
    ws2 = wb.create_sheet("说明")
    lines = [
        ["配置编辑器使用说明"],
        [""],
        ["1. 本 Excel 是「编辑层」，改完保存后运行生成器写回 JSON，系统读取 JSON。"],
        ["   运行: python scripts/config_excel_to_json.py"],
        [""],
        ["2. 「时间配置」sheet 列说明："],
        ["   - 配置名: 年度累计 / 月度数据 / 季度累计筛选 / 年基线数据（勿改）"],
        ["   - 模式: static（手填日期）或 dynamic（自动计算）"],
        ["   - 动态策略: 模式为 dynamic 时填写 last_full_month / last_full_quarter"],
        ["   - 开始日期/结束日期: 格式 YYYY-MM-DD 或 YYYY-MM-DD HH:MM:SS"],
        ["   - 年份/月份范围: 仅年基线数据使用，如 2025 / 1-8"],
        ["   - 说明: 填写配置用途（可选）"],
        [""],
        ["3. 「展示规则」sheet 列说明（路径列结构）："],
        ["   - 页面: 数据总览 / 年度达成 / 月度达成 / 季度达成 / 销售达成 / 年度同比"],
        ["   - 路径: 从页面下到配置项的路径，点号分隔"],
        ["     · 销售TopN                 → 页面下标量"],
        ["     · 客户矩阵.最大行数          → 区块.标量"],
        ["     · 部门卡.显示               → 区块.布尔（true/false）"],
        ["     · 客户矩阵.优先展示.1        → 区块.数组项（序号从1开始）"],
        ["     · 客户矩阵.客户筛选          → 空数组（值留空）"],
        ["   - 值: 配置值"],
        ["     · 数组项（优先展示 / 客户筛选）每个值一行，路径末尾带序号"],
        ["     · 显示类配置填 true / false"],
        ["     · 排序可选: 目标合计降序 / 实际金额降序 / 达成率降序"],
        ["   - 说明: 备注（可选）"],
        [""],
        ["4. 「销售归属」sheet 列说明："],
        ["   - 母公司: 客户归组的母公司名（如 广州小鹏汽车科技有限公司）"],
        ["   - 子公司: 实际结算主体名（=母公司时表示本部）"],
        ["   - 指标: 收入 / 回款 / 收入,回款（收入回款同构时用后者，一行搞定）"],
        ["   - 部门: 检测 / 信息 / 能源 / 海外"],
        ["   - 销售: 负责销售的姓名"],
        ["   - 比例: 分配比例，默认 1（全额）；拆分时填小数如 0.3 或百分比 30%"],
        ["   - 说明: 备注（可选）"],
        [""],
        ["5. 常见修改："],
        ["   - 换月份: 改「时间配置」的月度数据开始/结束日期"],
        ["   - 调整优先展示客户: 改「展示规则」对应页面的 客户矩阵.优先展示.N 行"],
        ["   - 显示全部客户: 把 客户矩阵.优先展示 的所有行删掉，保留一条空值行"],
        ["   - 改销售归属: 在「销售归属」sheet 直接改对应行的 销售 / 比例"],
        ["   - 拆分多销售: 把一行复制成多行，各填不同销售和比例（合计需=1）"],
        [""],
        ["6. 校验失败时生成器拒绝写回并提示原因，JSON 保持原样。"],
    ]
    for row in lines:
        ws2.append(row)
    ws2.column_dimensions["A"].width = 100
    ws2["A1"].font = Font(bold=True, size=14)

    wb.save(EXCEL_PATH)
    print(f"✅ 已生成模板: {EXCEL_PATH.relative_to(BASE_DIR)}")


# ──────────────────────────────────────────────────────────────
# 主入口
# ──────────────────────────────────────────────────────────────
def main():
    args = sys.argv[1:]
    init = "--init" in args
    dry_run = "--dry-run" in args

    if init:
        init_excel_template()
        return 0

    if not EXCEL_PATH.exists():
        print(f"未找到配置编辑器 Excel: {EXCEL_PATH.relative_to(BASE_DIR)}")
        print("请先运行: python scripts/config_excel_to_json.py --init")
        return 1

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)

    # 1) 时间配置 → cleaning_config.json
    if "时间配置" in wb.sheetnames:
        ws = wb["时间配置"]
        time_config = excel_to_time_config(ws)
        new_cfg = update_cleaning_config(time_config, dry_run=dry_run)
        if dry_run:
            print("【DRY-RUN】将写回以下时间范围：")
            print(json.dumps(new_cfg["时间范围"], ensure_ascii=False, indent=2))
        else:
            print("✅ 已更新 cleaning_config.json 的「时间范围」：")
            for k, v in time_config.items():
                print(f"  - {k}: {json.dumps(v, ensure_ascii=False)}")
    else:
        print("⚠️  Excel 缺少「时间配置」sheet，跳过")

    # 2) 展示规则 → 展示规则.json
    if "展示规则" in wb.sheetnames:
        ws_r = wb["展示规则"]
        rules = excel_to_display_rules(ws_r)
        new_rules = update_display_rules(rules, dry_run=dry_run)
        if dry_run:
            print("【DRY-RUN】将写回以下展示规则：")
            print(json.dumps(new_rules, ensure_ascii=False, indent=2))
        else:
            print("✅ 已更新 展示规则.json：")
            for page, obj in rules.items():
                keys = list(obj.keys())
                print(f"  - {page}: {keys}")
    else:
        print("⚠️  Excel 缺少「展示规则」sheet，跳过")

    # 3) 销售归属 → 客户销售归属.json
    if "销售归属" in wb.sheetnames:
        ws_a = wb["销售归属"]
        att = excel_to_attribution(ws_a)
        new_att = update_attribution(att, dry_run=dry_run)
        if dry_run:
            print("【DRY-RUN】将写回以下销售归属：")
            print(f"  母公司数: {len(new_att['客户归属'])}")
        else:
            n_subs = sum(len(g.get("子公司", {})) for g in att.values())
            print(f"✅ 已更新 客户销售归属.json：{len(att)} 母公司 / {n_subs} 子公司")
    else:
        print("⚠️  Excel 缺少「销售归属」sheet，跳过")

    return 0


if __name__ == "__main__":
    sys.exit(main())
